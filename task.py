from openpyxl import Workbook,load_workbook
import requests
from bs4 import BeautifulSoup

#Bu fonksiyon parametre olarak url alıp, o web sayfasındaki bilgileri array olarak döndürür.
def productInfo(url):
    #web sayfasına istek atılır, sonuç request.get methodu ile r değişkenine atanır.
    r = requests.get(url)
    #BeautifulSoup'un içine r.content değeri ve hangi kütüphane ile parse edileceği bilgisi verilir, böylelikle product değişkenine BeautifulSoup'tan çıkan lxml kodları döner.
    soup = BeautifulSoup(r.content,"lxml")
    #Sayfa kaynağındaki spesifik bir seçiciye sahip olan url adresi çekilmek istenilirse soup.find methodu kullanılır. Gereken bilgiler, sayfada "product__wrapper" altında olduğu için, buradaki özel değerler çekilir ve product değişkenine atanır. (Daha sonra daha da spesifik olan metrikler inceleneceği zaman yine incelenen alan daraltılacaktır)
    product = soup.find("div",attrs={"class":"product__wrapper"})
    #İncelemeler sonrasında mevcut olan ve olmayan bedenlerin tutulduğu lxml kodu bulunur ve sonrasında farklılıkları göz önünde bulundurularak bir ayrıştırma yapılır. Ardından bilgileri array olarak kaydolur.
    available = product.find_all("a",attrs={"class":"d-flex align-items-center justify-content-center text-reset product__variant product__size-variant mb-3 js-variant"})
    notAvailable = product.find_all("a",attrs={"class":"d-flex align-items-center justify-content-center text-reset product__variant product__size-variant mb-3 js-variant disabled"})
    #Eğer tüm bedenler var ise bölünme işleminin sorun yaratmaması için kontrol yapılır ve ona göre bir yüzde bulunur.
    if len(notAvailable) != 0:
        percentage = float(len(available)) / float(len(notAvailable))
    else:
        percentage = float(100)

    #İsim, marka, kod ve fiyat bilgilerinin tutulduğu lxml kodları incelenir. Gerekli bilgiler ise .text() yardımıyla string olarak, .strip() yadımıyla da boşluklardan temizlenmiş bir şekilde döndürülür. En son tüm istenilen bilgiler bir array olarak return değerine atanır.
    return [product.h1.text.strip(), product.find("a", attrs={"class":"product__brand"}).text.strip(), product.find("div", attrs={"class":"product__code d-block d-lg-none"}).text.strip(), product.find("div",attrs={"class":"product__price"}).text.strip(), percentage]




#Çalışılacak excel dosyası açılıp aktive edilmiş ve ardından çalışalacak sheet belirlenmiştir. (Mailde gelen dosya adı boşluklu olduğu için sorun çıkmaması adına ismi daha önceden değiştirilmiştir. Yazılan bu python dosyası terminalden çalıştırılmaktadır ve de MacOS işletim sistemi için boşlukların sık sık sorun olduğu gözlemlenmiştir.)
workBook = load_workbook("productDetailURLFile.xlsx")
workSheet = workBook.active
workSheet = workBook["Sheet4"]

#Gidilecek urllerin tutulabilmeleri ve de karşılığında alınan bilgilerin tutulması için için boş iki array açılmıştır.
urlArray =  []
productInfoList = []

#Urllerin bulunabilmesi için hepsinin önüne "https://www.spx.com.tr" eklenmiştir ve de array excel boyutu kadar(100) doldurulmuştur. Excel satır ve sutünlarının 1den başladığı da hesaba katılmıştır.
for item in range(1,101):
    s = "https://www.spx.com.tr"
    s= s+workSheet.cell(item,1).value
    urlArray.append(s)

#Her urlnin denk geldiği ürünün bilgileri tanımlanan productInfo fonksiyonu sayesinde bir arraye dönüşebilmektedir. Her ürünün array'i productInfoList'de tutulmaktadır.
for item in urlArray:
    productInfoList.append(productInfo(item))

#productInfoList tamamanlanınca, alınan tüm bilgiler productDetailURLFile excel dosyasına geri yazılmaktadır. Yeni bir excel dosyası açılmış ve oraya yazılmaları da denenmiştir ancak bu assignmentdaki senaryoda, müşterilere bilgi sunulacağı için istenilen urlerin yanında bilgilerin girilmesinin daha sistematik olacağı düşünülmüştür.
for item in range(len(productInfoList)):
    for data in range(len(productInfoList[item])):
        workSheet.cell(row = item+1, column = data+2).value = productInfoList[item][data]

#En son ise karmaşıklığı azaltmak adına, girilen ürün bilgilerinin altına hangi metriğe dahil oldukları gösterilmiştir. En yukarı yazılamamıştır çünkü gönderilen excel dosyasındaki urller birinci satırdan başlamaktadır.
workSheet.cell(row = 101, column = 2).value = "Product Name"
workSheet.cell(row = 101, column = 3).value = "Product Brand"
workSheet.cell(row = 101, column = 4).value = "Product Code"
workSheet.cell(row = 101, column = 5).value = "Product Price"
workSheet.cell(row = 101, column = 6).value = "Product Availability"

#Üzerinde çalışılan excel dosyası kaydedilip kapanmıştır.
workBook.save("productDetailURLFile.xlsx")
workBook.close()
